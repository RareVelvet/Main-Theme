#!/usr/bin/env python3
"""
build_tracker.py — Intentional Budget Tracker (Excel edition)

Generates "Intentional Budget Tracker.xlsx": a multi-tab, formula-driven budget
workbook. The buyer types their own numbers on "Start Here" and every derived
figure (Freedom Number, savings rate, net worth, timeline, debt payoff)
recalculates live via native Excel formulas. No macros — pure .xlsx, so it opens
in Excel, LibreOffice, and Google Sheets.

Run:  python build_tracker.py
Out:  Intentional Budget Tracker.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter, quote_sheetname

# --------------------------------------------------------------------------- #
# Brand palette (hex, no '#')
# --------------------------------------------------------------------------- #
INK       = "3A342B"   # primary text
MUTED     = "8C8475"   # secondary text
LINE      = "E8E0D3"   # gridline / border
CANVAS    = "FAF7F1"   # sheet background tint
SAGE      = "C2D2B9"   # investments / positive
ROSE      = "E3C7BD"   # guilt-free / over-budget
BLUE      = "BCC8D8"   # fixed costs
HONEY     = "E0C994"   # freedom number / accents
GREEN_POS = "2E7D4F"   # on-plan / under-budget text
RED_NEG   = "C0392B"   # over-budget / over-allocated text
HONEY_LT  = "F2E6C9"   # honey input-cell tint
WHITE     = "FFFFFF"

# Currency is parameterized so we can emit one workbook per currency.
# CURRENCY_SYMBOL / CURRENCY_FMT are reassigned per-variant in main().
CURRENCY_SYMBOL = "$"
CURRENCY_FMT    = '"$"#,##0'
PCT_FMT      = '0.0%'
YEAR_FMT     = '0'

# (code, symbol, number format) — symbol drives the Start Here dropdown default,
# the format drives every currency cell. Scandinavian/Swiss put the symbol after.
CURRENCIES = [
    ("USD", "$",   '"$"#,##0'),
    ("GBP", "£",   '"£"#,##0'),
    ("EUR", "€",   '"€"#,##0'),
    ("JPY", "¥",   '"¥"#,##0'),
    ("INR", "₹",   '"₹"#,##0'),
    ("BRL", "R$",  '"R$"#,##0'),
    ("SEK", "kr",  '#,##0" kr"'),
    ("CHF", "CHF", '"CHF "#,##0'),
]

# --------------------------------------------------------------------------- #
# Reusable style helpers
# --------------------------------------------------------------------------- #
HEAD_FONT   = Font(name="Georgia", size=14, bold=True, color=INK)
TITLE_FONT  = Font(name="Georgia", size=18, bold=True, color=INK)
SUB_FONT    = Font(name="Georgia", size=11, italic=True, color=MUTED)
SECTION_FONT = Font(name="Georgia", size=11, bold=True, color=INK)
LABEL_FONT  = Font(name="Calibri", size=11, color=INK)
MUTED_FONT  = Font(name="Calibri", size=10, color=MUTED)
BOLD_FONT   = Font(name="Calibri", size=11, bold=True, color=INK)
COLHEAD_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BIGNUM_FONT = Font(name="Consolas", size=16, bold=True, color=INK)

thin = Side(style="thin", color=LINE)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_BOTTOM = Border(bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")
RIGHT  = Alignment(horizontal="right", vertical="center")
WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)

INPUT_FILL   = PatternFill("solid", fgColor=HONEY_LT)
HONEY_FILL   = PatternFill("solid", fgColor=HONEY)
SAGE_FILL    = PatternFill("solid", fgColor=SAGE)
ROSE_FILL    = PatternFill("solid", fgColor=ROSE)
BLUE_FILL    = PatternFill("solid", fgColor=BLUE)
LINE_FILL    = PatternFill("solid", fgColor=LINE)
CANVAS_FILL  = PatternFill("solid", fgColor=CANVAS)
INK_FILL     = PatternFill("solid", fgColor=INK)
UNLOCK = Protection(locked=False)


def style_cell(ws, coord, value=None, font=LABEL_FONT, fill=None, align=LEFT,
               number_format=None, border=None, protection=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = align
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    if protection:
        c.protection = protection
    return c


def title_block(ws, title, subtitle, span="A1:H1"):
    """Title row + subtitle row, merged across `span` columns."""
    first_col = span.split(":")[0]
    last_col = span.split(":")[1].rstrip("0123456789")
    ws.merge_cells(f"{first_col[0]}1:{last_col}1")
    ws.merge_cells(f"{first_col[0]}2:{last_col}2")
    style_cell(ws, f"{first_col[0]}1", title, font=TITLE_FONT, align=LEFT)
    style_cell(ws, f"{first_col[0]}2", subtitle, font=SUB_FONT, align=LEFT)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18


def protect(ws, allow_insert=False):
    """Enable sheet protection (no password). Locked cells can't be edited;
    unlocked (honey) input cells can. Optionally allow row insertion."""
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False     # navigate but skip locked cells
    ws.protection.selectUnlockedCells = True
    ws.protection.formatCells = False
    ws.protection.insertRows = allow_insert
    ws.protection.sort = True
    ws.protection.autoFilter = True


def section_header(ws, row, text, fill, span_cols):
    """A full-width section band row, e.g. 'FIXED COSTS'."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT
    c.fill = fill
    c.alignment = LEFT
    ws.row_dimensions[row].height = 20
    for col in range(1, span_cols + 1):
        ws.cell(row=row, column=col).fill = fill


# =========================================================================== #
#  SHEET: Read Me
# =========================================================================== #
def build_readme(wb):
    ws = wb.create_sheet("Read Me")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    title_block(ws, "Intentional Budget Tracker",
                "A calm, formula-driven money workbook — type your numbers, watch your future update.",
                span="B1:B1")

    lines = [
        ("h", "How to use this workbook"),
        ("p", "1.  Open the “Start Here” tab and replace every honey-shaded cell with your own numbers."),
        ("p", "2.  Everything else recalculates automatically — Freedom Number, savings rate, net worth, timeline."),
        ("p", "3.  Work through the tabs left to right. Only the honey-shaded cells are meant to be edited."),
        ("p", "4.  Once a month, update “Monthly Spending” actuals and add a row to “Net Worth Log”."),
        ("sp", ""),
        ("h", "What the tabs do"),
        ("p", "• Start Here — all your inputs + your headline numbers."),
        ("p", "• Monthly Spending — your Conscious Spending Plan (budget vs. actual, savings rate)."),
        ("p", "• Net Worth — assets minus liabilities, and where your money lives."),
        ("p", "• Net Worth Log — one row per month; the chart shows your growth."),
        ("p", "• Freedom Number — the number that means work becomes optional."),
        ("p", "• Freedom Timeline — a 30-year projection and your work-optional year."),
        ("p", "• Debts — payoff months, interest, and what an extra $200/mo saves you."),
        ("p", "• Annual Vision / Reviews / Reflection — space to think, not just track."),
        ("sp", ""),
        ("h", "Good habits"),
        ("p", "• Use NET (take-home) income for budgeting, GROSS only for context."),
        ("p", "• Budget every dollar: aim to get “Left to assign” on Monthly Spending to $0."),
        ("p", "• Update actuals weekly so the month never surprises you."),
        ("p", "• Log net worth the same day each month — consistency beats precision."),
        ("p", "• Be honest in Guilt-Free Spending; the goal is intention, not guilt."),
        ("sp", ""),
        ("muted", "Tip: cells with formulas are locked so you can’t break them by accident. "
                  "Only the honey-shaded inputs are editable. To unlock everything, choose "
                  "Review ▸ Unprotect Sheet (there is no password)."),
    ]
    r = 4
    for kind, text in lines:
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "h":
            cell.font = HEAD_FONT
            ws.row_dimensions[r].height = 24
        elif kind == "muted":
            cell.font = MUTED_FONT
            cell.alignment = WRAP
            ws.row_dimensions[r].height = 46
        elif kind == "sp":
            ws.row_dimensions[r].height = 6
        else:
            cell.font = LABEL_FONT
            cell.alignment = WRAP
            ws.row_dimensions[r].height = 18
        r += 1
    return ws


# =========================================================================== #
#  SHEET: Start Here  (inputs + named ranges + headline)
# =========================================================================== #
def build_start_here(wb):
    ws = wb.create_sheet("Start Here")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 34
    ws.column_dimensions["F"].width = 18
    title_block(ws, "Start Here",
                "Type your numbers in the honey cells. Everything else updates itself.",
                span="B1:F1")

    # ---- Inputs --------------------------------------------------------- #
    section_band(ws, 4, "YOUR INPUTS", HONEY_FILL, 2, 3)
    inputs = [
        # (row, label, value, named_range, number_format)
        (5,  "Your name",                     "Friend", "Name",          None),
        (6,  "Currency symbol",        CURRENCY_SYMBOL, "Currency",      None),
        (7,  "Planning year",                 2026,     "Year",          YEAR_FMT),
        (8,  "Your age",                       38,      "Age",           '0'),
        (9,  "Monthly gross income",           9500,    "IncomeGross",   CURRENCY_FMT),
        (10, "Monthly net (take-home)",        6800,    "IncomeNet",     CURRENCY_FMT),
        (11, "Ideal annual expenses",          48000,   "AnnualExpenses",CURRENCY_FMT),
        (12, "Passive income (net, annual)",   6000,    "PassiveIncome", CURRENCY_FMT),
        (13, "Multiplier (25/28/30/33)",       25,      "Multiplier",    '0'),
        (14, "Expected real return (%)",       6,       "ReturnRate",    '0.0'),
        (15, "Monthly invested",               1200,    "MonthlyInvested",CURRENCY_FMT),
        (16, "Breathing monthly income",       3200,    "BreatheMonthly",CURRENCY_FMT),
    ]
    named_cells = {}
    for row, label, value, name, fmt in inputs:
        style_cell(ws, f"B{row}", label, font=LABEL_FONT, align=LEFT)
        c = style_cell(ws, f"C{row}", value, font=BOLD_FONT, fill=INPUT_FILL,
                       align=RIGHT, number_format=fmt, border=BORDER_ALL,
                       protection=UNLOCK)
        named_cells[name] = f"'Start Here'!$C${row}"

    # ---- Headline numbers ----------------------------------------------- #
    section_band(ws, 4, "YOUR NUMBERS", HONEY_FILL, 5, 6)
    headline = [
        (5,  "Freedom Number",            "=(AnnualExpenses-PassiveIncome)*Multiplier", CURRENCY_FMT),
        (7,  "Monthly freedom income",    "=AnnualExpenses/12",                          CURRENCY_FMT),
        (9,  "Breathing Number (monthly)","=BreatheMonthly",                             CURRENCY_FMT),
        (11, "Breathing asset base",      "=BreatheMonthly*12*Multiplier",               CURRENCY_FMT),
        (13, "Work-optional year",
             "=IFERROR(Year+CEILING(NPER(ReturnRate/100/12,-MonthlyInvested,-NetWorth,FreedomNumber)/12,1),Year)",
             '0'),
    ]
    for row, label, formula, fmt in headline:
        style_cell(ws, f"E{row}", label, font=LABEL_FONT, align=LEFT)
        style_cell(ws, f"F{row}", formula, font=BIGNUM_FONT, fill=CANVAS_FILL,
                   align=RIGHT, number_format=fmt, border=BORDER_ALL)
        ws.row_dimensions[row].height = 26
    # FreedomNumber named range -> headline cell
    named_cells["FreedomNumber"] = "'Start Here'!$F$5"

    # Multiplier dropdown
    dv = DataValidation(type="list", formula1='"25,28,30,33"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["C13"])
    # Currency dropdown
    dvc = DataValidation(type="list", formula1='"$,£,€,¥,₹,R$,kr,CHF"', allow_blank=False)
    ws.add_data_validation(dvc)
    dvc.add(ws["C6"])

    note = ("Multiplier rule of thumb: 25 = the classic 4% rule, 33 ≈ a safer 3%. "
            "“Real return” already subtracts inflation — 6% is a reasonable long-run stock/bond blend.")
    ws.merge_cells("B18:F19")
    style_cell(ws, "B18", note, font=MUTED_FONT, align=WRAP)
    ws.row_dimensions[18].height = 30

    protect(ws)
    return ws, named_cells


def section_band(ws, row, text, fill, col_start, col_end):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = fill
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = SECTION_FONT
    c.alignment = LEFT
    ws.row_dimensions[row].height = 20


# =========================================================================== #
#  SHEET: Monthly Spending
# =========================================================================== #
def build_monthly_spending(wb):
    ws = wb.create_sheet("Monthly Spending")
    ws.sheet_view.showGridLines = False
    widths = [22, 13, 13, 13, 11, 13, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    title_block(ws, "Monthly Spending",
                "Your Conscious Spending Plan — give every dollar a job.", span="A1:G1")

    headers = ["Category", "Budget", "Actual", "Variance", "% of Net", "Status", "Due day"]
    hr = 3
    for i, h in enumerate(headers, start=1):
        style_cell(ws, f"{get_column_letter(i)}{hr}", h, font=COLHEAD_FONT,
                   fill=INK_FILL, align=CENTER, border=BORDER_ALL)
    ws.freeze_panes = "A4"

    sections = [
        ("FIXED COSTS", BLUE_FILL, "fixed", [
            ("Rent / Mortgage", 1650, 1650, 1),
            ("Utilities", 145, 162, 5),
            ("Health insurance", 285, 285, 1),
            ("Groceries", 620, 648, 0),
            ("Car payment", 380, 380, 10),
            ("Subscriptions", 85, 95, 15),
        ]),
        ("INVESTMENTS", SAGE_FILL, "invest", [
            ("401(k) contribution", 700, 700, 1),
            ("Roth IRA", 500, 500, 1),
        ]),
        ("SAVINGS GOALS", HONEY_FILL, "savings", [
            ("Emergency fund", 200, 200, 1),
            ("Vacation fund", 150, 150, 1),
        ]),
        ("GUILT-FREE SPENDING", ROSE_FILL, "guilt", [
            ("Dining out", 380, 428, 0),
            ("Joy / Fun", 280, 295, 0),
            ("Giving", 350, 350, 0),
        ]),
    ]

    row = 4
    subtotal_actual = {}   # kind -> "C{row}"
    section_data_rows = []  # (first, last) item-row spans for CF
    for name, fill, kind, items in sections:
        section_header(ws, row, name, fill, 7)
        row += 1
        first = row
        for label, budget, actual, due in items:
            style_cell(ws, f"A{row}", label, font=LABEL_FONT, border=BORDER_BOTTOM)
            style_cell(ws, f"B{row}", budget, font=BOLD_FONT, fill=INPUT_FILL,
                       align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL,
                       protection=UNLOCK)
            style_cell(ws, f"C{row}", actual, font=BOLD_FONT, fill=INPUT_FILL,
                       align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL,
                       protection=UNLOCK)
            style_cell(ws, f"D{row}", f"=C{row}-B{row}", align=RIGHT,
                       number_format=CURRENCY_FMT, border=BORDER_BOTTOM)
            style_cell(ws, f"E{row}", f"=IF(IncomeNet=0,0,C{row}/IncomeNet)",
                       align=RIGHT, number_format=PCT_FMT, border=BORDER_BOTTOM)
            if kind == "savings":
                status = f'=IF(C{row}>=B{row},"On goal ✓","Short")'
            else:
                status = f'=IF(C{row}>B{row},"Over","On plan")'
            style_cell(ws, f"F{row}", status, align=CENTER, border=BORDER_BOTTOM)
            style_cell(ws, f"G{row}", due, fill=INPUT_FILL, align=CENTER,
                       number_format='0', border=BORDER_ALL, protection=UNLOCK)
            row += 1
        last = row - 1
        section_data_rows.append((first, last))
        # subtotal row
        style_cell(ws, f"A{row}", f"  Subtotal — {name.title()}", font=BOLD_FONT)
        style_cell(ws, f"B{row}", f"=SUM(B{first}:B{last})", font=BOLD_FONT,
                   align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)
        style_cell(ws, f"C{row}", f"=SUM(C{first}:C{last})", font=BOLD_FONT,
                   align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)
        style_cell(ws, f"D{row}", f"=C{row}-B{row}", font=BOLD_FONT, align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL)
        style_cell(ws, f"E{row}", f"=IF(IncomeNet=0,0,C{row}/IncomeNet)",
                   font=BOLD_FONT, align=RIGHT, number_format=PCT_FMT, border=BORDER_ALL)
        subtotal_actual[kind] = f"C{row}"
        row += 2

    # ---- Summary block --------------------------------------------------- #
    sum_row = row + 1
    all_actuals = "+".join(subtotal_actual.values())
    style_cell(ws, f"A{sum_row}", "Left to assign", font=HEAD_FONT)
    lta = style_cell(ws, f"C{sum_row}", f"=IncomeNet-({all_actuals})",
                     font=BIGNUM_FONT, align=RIGHT, number_format=CURRENCY_FMT,
                     border=BORDER_ALL)
    style_cell(ws, f"D{sum_row}",
               '=IF(C%d<0,"Over-allocated",IF(C%d=0,"Every dollar assigned ✓","Unassigned"))'
               % (sum_row, sum_row), font=MUTED_FONT, align=LEFT)

    sr_row = sum_row + 1
    style_cell(ws, f"A{sr_row}", "Savings rate", font=HEAD_FONT)
    style_cell(ws, f"C{sr_row}",
               f"=IF(IncomeNet=0,0,({subtotal_actual['invest']}+{subtotal_actual['savings']})/IncomeNet)",
               font=BIGNUM_FONT, align=RIGHT, number_format=PCT_FMT, border=BORDER_ALL)
    style_cell(ws, f"D{sr_row}", "Investments + savings ÷ take-home", font=MUTED_FONT)

    # ---- Conditional formatting ----------------------------------------- #
    red_font = Font(name="Calibri", color=RED_NEG, bold=True)
    green_font = Font(name="Calibri", color=GREEN_POS, bold=True)
    for first, last in section_data_rows:
        rng = f"D{first}:D{last}"
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="lessThan", formula=["0"], font=red_font))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="greaterThan", formula=["0"], font=green_font))
        # data bars on % of Net
        ws.conditional_formatting.add(f"E{first}:E{last}",
            DataBarRule(start_type="num", start_value=0, end_type="num",
                        end_value=0.5, color=BLUE))
    # Left to assign -> rose fill if negative
    ws.conditional_formatting.add(f"C{sum_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=ROSE_FILL))

    protect(ws)
    return ws


# =========================================================================== #
#  SHEET: Net Worth
# =========================================================================== #
def build_net_worth(wb):
    ws = wb.create_sheet("Net Worth")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [26, 16, 18, 3, 20, 16]):
        ws.column_dimensions[col].width = w
    title_block(ws, "Net Worth", "What you own, minus what you owe.", span="A1:F1")

    locations = "Taxable,Tax-advantaged,Real estate,Business,Cash,Other"
    dv = DataValidation(type="list", formula1=f'"{locations}"', allow_blank=True)
    ws.add_data_validation(dv)

    # ---- Assets ---------------------------------------------------------- #
    section_band(ws, 3, "ASSETS", SAGE_FILL, 1, 3)
    for i, h in enumerate(["Asset", "Value", "Location"], start=1):
        style_cell(ws, f"{get_column_letter(i)}4", h, font=COLHEAD_FONT,
                   fill=INK_FILL, align=CENTER, border=BORDER_ALL)
    assets = [
        ("401(k)", 45000, "Tax-advantaged"),
        ("Roth IRA", 22000, "Tax-advantaged"),
        ("Brokerage", 18000, "Taxable"),
        ("Home (market value)", 320000, "Real estate"),
        ("Checking", 6000, "Cash"),
        ("Emergency savings", 12000, "Cash"),
        ("", None, ""),
        ("", None, ""),
    ]
    a_first = 5
    for j, (name, val, loc) in enumerate(assets):
        r = a_first + j
        style_cell(ws, f"A{r}", name, fill=INPUT_FILL, border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"B{r}", val, fill=INPUT_FILL, align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"C{r}", loc, fill=INPUT_FILL, align=CENTER,
                   border=BORDER_ALL, protection=UNLOCK)
        dv.add(ws[f"C{r}"])
    a_last = a_first + len(assets) - 1
    ta_row = a_last + 1
    style_cell(ws, f"A{ta_row}", "Total assets", font=BOLD_FONT)
    style_cell(ws, f"B{ta_row}", f"=SUM(B{a_first}:B{a_last})", font=BOLD_FONT,
               align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL, fill=SAGE_FILL)

    # ---- Liabilities ----------------------------------------------------- #
    l_head = ta_row + 2
    section_band(ws, l_head, "LIABILITIES", ROSE_FILL, 1, 3)
    for i, h in enumerate(["Liability", "Value"], start=1):
        style_cell(ws, f"{get_column_letter(i)}{l_head+1}", h, font=COLHEAD_FONT,
                   fill=INK_FILL, align=CENTER, border=BORDER_ALL)
    liabilities = [
        ("Mortgage", 240000),
        ("Car loan", 9000),
        ("Credit card", 1500),
        ("", None),
        ("", None),
    ]
    l_first = l_head + 2
    for j, (name, val) in enumerate(liabilities):
        r = l_first + j
        style_cell(ws, f"A{r}", name, fill=INPUT_FILL, border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"B{r}", val, fill=INPUT_FILL, align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL, protection=UNLOCK)
    l_last = l_first + len(liabilities) - 1
    tl_row = l_last + 1
    style_cell(ws, f"A{tl_row}", "Total liabilities", font=BOLD_FONT)
    style_cell(ws, f"B{tl_row}", f"=SUM(B{l_first}:B{l_last})", font=BOLD_FONT,
               align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL, fill=ROSE_FILL)

    # ---- Net worth ------------------------------------------------------- #
    nw_row = tl_row + 2
    style_cell(ws, f"A{nw_row}", "NET WORTH", font=HEAD_FONT)
    nw_cell = style_cell(ws, f"B{nw_row}", f"=B{ta_row}-B{tl_row}", font=BIGNUM_FONT,
                         align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL,
                         fill=HONEY_FILL)

    # ---- Where your money lives (SUMIF by location) ---------------------- #
    style_cell(ws, "E3", "WHERE YOUR MONEY LIVES", font=SECTION_FONT, fill=HONEY_FILL)
    ws.merge_cells("E3:F3")
    for i, h in enumerate(["Location", "Total"], start=5):
        style_cell(ws, f"{get_column_letter(i)}4", h, font=COLHEAD_FONT,
                   fill=INK_FILL, align=CENTER, border=BORDER_ALL)
    loc_list = locations.split(",")
    for k, loc in enumerate(loc_list):
        r = 5 + k
        style_cell(ws, f"E{r}", loc, border=BORDER_ALL)
        style_cell(ws, f"F{r}",
                   f'=SUMIF($C${a_first}:$C${a_last},E{r},$B${a_first}:$B${a_last})',
                   align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)

    protect(ws, allow_insert=True)
    return ws, f"'Net Worth'!$B${nw_row}"


# =========================================================================== #
#  SHEET: Net Worth Log
# =========================================================================== #
def build_net_worth_log(wb, nw_ref):
    ws = wb.create_sheet("Net Worth Log")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    title_block(ws, "Net Worth Log",
                "Add one row each month. The chart tracks your growth.", span="A1:E1")

    for i, h in enumerate(["Month", "Net worth"], start=1):
        style_cell(ws, f"{get_column_letter(i)}3", h, font=COLHEAD_FONT,
                   fill=INK_FILL, align=CENTER, border=BORDER_ALL)
    ws.freeze_panes = "A4"

    # Seed a few months; first row pulls live net worth so there's always data.
    seed = [
        ("2026-01", 178000),
        ("2026-02", 181500),
        ("2026-03", 184200),
        ("2026-04", 187900),
        ("2026-05", 190600),
        ("2026-06", f"={nw_ref}"),
    ]
    first = 4
    for j, (m, v) in enumerate(seed):
        r = first + j
        style_cell(ws, f"A{r}", m, fill=INPUT_FILL, align=CENTER,
                   border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"B{r}", v, fill=INPUT_FILL, align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL, protection=UNLOCK)
    last = first + 30  # leave empty rows for future logging
    for r in range(first + len(seed), last + 1):
        style_cell(ws, f"A{r}", None, fill=INPUT_FILL, align=CENTER,
                   border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"B{r}", None, fill=INPUT_FILL, align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL, protection=UNLOCK)

    # Line chart
    chart = LineChart()
    chart.title = "Net worth over time"
    chart.style = 2
    chart.height = 8
    chart.width = 16
    data = Reference(ws, min_col=2, min_row=3, max_row=last)
    cats = Reference(ws, min_col=1, min_row=4, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.numFmt = CURRENCY_FMT
    chart.y_axis.majorGridlines = None
    ws.add_chart(chart, "D4")

    protect(ws, allow_insert=True)
    return ws


# =========================================================================== #
#  SHEET: Freedom Number
# =========================================================================== #
def build_freedom_number(wb):
    ws = wb.create_sheet("Freedom Number")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    title_block(ws, "Freedom Number",
                "The amount that makes paid work optional.", span="B1:C1")

    rows = [
        ("Covered expenses (annual)", "=AnnualExpenses-PassiveIncome", CURRENCY_FMT),
        ("Freedom Number", "=(AnnualExpenses-PassiveIncome)*Multiplier", CURRENCY_FMT),
        ("Your net worth today", "=NetWorth", CURRENCY_FMT),
        ("Progress to freedom", "=IF(FreedomNumber=0,0,NetWorth/FreedomNumber)", PCT_FMT),
        ("Gap remaining", "=MAX(FreedomNumber-NetWorth,0)", CURRENCY_FMT),
        ("Breathing asset base", "=BreatheMonthly*12*Multiplier", CURRENCY_FMT),
        ("Breathing progress",
         "=IF(BreatheMonthly=0,0,NetWorth/(BreatheMonthly*12*Multiplier))", PCT_FMT),
    ]
    r = 4
    progress_cell = None
    for label, formula, fmt in rows:
        style_cell(ws, f"B{r}", label, font=LABEL_FONT)
        c = style_cell(ws, f"C{r}", formula, font=BIGNUM_FONT, fill=CANVAS_FILL,
                       align=RIGHT, number_format=fmt, border=BORDER_ALL)
        if label == "Progress to freedom":
            progress_cell = f"C{r}"
        ws.row_dimensions[r].height = 24
        r += 1

    if progress_cell:
        ws.conditional_formatting.add(progress_cell,
            DataBarRule(start_type="num", start_value=0, end_type="num",
                        end_value=1, color=HONEY))

    # Reflection questions
    qr = r + 1
    style_cell(ws, f"B{qr}", "Five questions worth sitting with", font=HEAD_FONT)
    qr += 1
    questions = [
        "1.  What would you do with your days if work were optional?",
        "2.  What does “enough” actually look like for your life?",
        "3.  Which expenses buy real joy — and which are just habit?",
        "4.  What could you do now that your future self would thank you for?",
        "5.  Who do you want to become on the way to this number?",
    ]
    for q in questions:
        style_cell(ws, f"B{qr}", q, font=LABEL_FONT, align=WRAP)
        ws.merge_cells(f"B{qr}:C{qr}")
        ws.row_dimensions[qr].height = 20
        qr += 1

    protect(ws)
    return ws


# =========================================================================== #
#  SHEET: Freedom Timeline
# =========================================================================== #
def build_freedom_timeline(wb):
    ws = wb.create_sheet("Freedom Timeline")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [8, 7, 16, 15, 14, 16]):
        ws.column_dimensions[col].width = w
    title_block(ws, "Freedom Timeline",
                "A 30-year projection. The honey row is the year work becomes optional.",
                span="A1:F1")

    headers = ["Year", "Age", "Start balance", "Contributions", "Growth", "End balance"]
    hr = 3
    for i, h in enumerate(headers, start=1):
        style_cell(ws, f"{get_column_letter(i)}{hr}", h, font=COLHEAD_FONT,
                   fill=INK_FILL, align=CENTER, border=BORDER_ALL)
    ws.freeze_panes = "A4"

    first = 4
    n_years = 30
    for k in range(n_years):
        r = first + k
        if k == 0:
            style_cell(ws, f"A{r}", "=Year", align=CENTER, number_format=YEAR_FMT, border=BORDER_ALL)
            style_cell(ws, f"B{r}", "=Age", align=CENTER, number_format='0', border=BORDER_ALL)
            style_cell(ws, f"C{r}", "=NetWorth", align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)
        else:
            style_cell(ws, f"A{r}", f"=A{r-1}+1", align=CENTER, number_format=YEAR_FMT, border=BORDER_ALL)
            style_cell(ws, f"B{r}", f"=B{r-1}+1", align=CENTER, number_format='0', border=BORDER_ALL)
            style_cell(ws, f"C{r}", f"=F{r-1}", align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)
        style_cell(ws, f"D{r}", "=MonthlyInvested*12", align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL)
        style_cell(ws, f"F{r}", f"=FV(ReturnRate/100/12,12,-MonthlyInvested,-C{r})",
                   align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)
        style_cell(ws, f"E{r}", f"=F{r}-C{r}-D{r}", align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL)
    last = first + n_years - 1

    # Conditional format: honey fill on the first row crossing the Freedom Number
    ws.conditional_formatting.add(f"A{first}:F{last}",
        FormulaRule(formula=[f"AND($F{first}>=FreedomNumber,$C{first}<FreedomNumber)"],
                    fill=HONEY_FILL, stopIfTrue=False))

    # Line chart: End balance vs Year + Freedom Number reference line
    chart = LineChart()
    chart.title = "Projected balance vs. Freedom Number"
    chart.height = 9
    chart.width = 18
    data = Reference(ws, min_col=6, min_row=3, max_row=last)   # End balance
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Build a helper column H with the (constant) Freedom Number for the ref line
    style_cell(ws, "H3", "Freedom Number", font=COLHEAD_FONT, fill=INK_FILL, align=CENTER)
    for k in range(n_years):
        r = first + k
        style_cell(ws, f"H{r}", "=FreedomNumber", number_format=CURRENCY_FMT)
    ws.column_dimensions["H"].width = 16
    ref_data = Reference(ws, min_col=8, min_row=3, max_row=last)
    chart.add_data(ref_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.numFmt = CURRENCY_FMT
    ws.add_chart(chart, "A36")

    # ---- What-if panel --------------------------------------------------- #
    wr = 4
    style_cell(ws, "J3", "WHAT IF…", font=SECTION_FONT, fill=HONEY_FILL)
    ws.merge_cells("J3:L3")
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 16
    whatifs = [
        ("Extra monthly invested", 300, CURRENCY_FMT, "extra"),
        ("Higher annual return (%)", 7, '0.0', "ret"),
        ("One-time lump sum now", 10000, CURRENCY_FMT, "lump"),
    ]
    cellmap = {}
    for j, (label, val, fmt, key) in enumerate(whatifs):
        r = wr + j
        style_cell(ws, f"J{r}", label, font=LABEL_FONT)
        style_cell(ws, f"K{r}", val, font=BOLD_FONT, fill=INPUT_FILL, align=RIGHT,
                   number_format=fmt, border=BORDER_ALL, protection=UNLOCK)
        cellmap[key] = f"K{r}"
    new_year_row = wr + len(whatifs) + 1
    style_cell(ws, f"J{new_year_row}", "New work-optional year", font=HEAD_FONT)
    formula = (f"=IFERROR(Year+CEILING(NPER({cellmap['ret']}/100/12,"
               f"-(MonthlyInvested+{cellmap['extra']}),"
               f"-(NetWorth+{cellmap['lump']}),FreedomNumber)/12,1),Year)")
    style_cell(ws, f"K{new_year_row}", formula, font=BIGNUM_FONT, fill=CANVAS_FILL,
               align=RIGHT, number_format='0', border=BORDER_ALL)

    protect(ws)
    return ws


# =========================================================================== #
#  SHEET: Debts
# =========================================================================== #
def build_debts(wb):
    ws = wb.create_sheet("Debts")
    ws.sheet_view.showGridLines = False
    headers = ["Debt", "Balance", "APR %", "Payment", "Months to payoff",
               "Interest remaining", "Payoff year", "Months (+$200)", "Interest saved (+$200)"]
    widths = [18, 14, 9, 12, 16, 18, 12, 15, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    title_block(ws, "Debts", "Payoff timelines — and what an extra $200/mo buys you.",
                span="A1:I1")

    hr = 3
    for i, h in enumerate(headers, start=1):
        style_cell(ws, f"{get_column_letter(i)}{hr}", h, font=COLHEAD_FONT,
                   fill=INK_FILL, align=CENTER, border=BORDER_ALL)
    ws.freeze_panes = "A4"

    debts = [
        ("Credit card", 1500, 22, 150),
        ("Car loan", 9000, 6, 380),
        ("Student loan", 12000, 5, 200),
        ("", None, None, None),
        ("", None, None, None),
    ]
    first = 4
    for j, (name, bal, apr, pay) in enumerate(debts):
        r = first + j
        style_cell(ws, f"A{r}", name, fill=INPUT_FILL, border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"B{r}", bal, fill=INPUT_FILL, align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"C{r}", apr, fill=INPUT_FILL, align=RIGHT,
                   number_format='0.0', border=BORDER_ALL, protection=UNLOCK)
        style_cell(ws, f"D{r}", pay, fill=INPUT_FILL, align=RIGHT,
                   number_format=CURRENCY_FMT, border=BORDER_ALL, protection=UNLOCK)
        # Months to payoff
        style_cell(ws, f"E{r}",
                   f'=IFERROR(ROUNDUP(NPER(C{r}/100/12,-D{r},B{r}),0),"—")',
                   align=RIGHT, number_format='0', border=BORDER_ALL)
        # Interest remaining
        style_cell(ws, f"F{r}",
                   f'=IF(ISNUMBER(E{r}),D{r}*E{r}-B{r},"—")',
                   align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)
        # Payoff year
        style_cell(ws, f"G{r}",
                   f'=IF(ISNUMBER(E{r}),Year+ROUNDUP(E{r}/12,0),"—")',
                   align=RIGHT, number_format=YEAR_FMT, border=BORDER_ALL)
        # Months with +$200
        style_cell(ws, f"H{r}",
                   f'=IFERROR(ROUNDUP(NPER(C{r}/100/12,-(D{r}+200),B{r}),0),"—")',
                   align=RIGHT, number_format='0', border=BORDER_ALL)
        # Interest saved with +$200
        style_cell(ws, f"I{r}",
                   f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(H{r})),F{r}-((D{r}+200)*H{r}-B{r}),"—")',
                   align=RIGHT, number_format=CURRENCY_FMT, border=BORDER_ALL)
    last = first + len(debts) - 1

    tr = last + 1
    style_cell(ws, f"A{tr}", "Total debt", font=BOLD_FONT)
    style_cell(ws, f"B{tr}", f"=SUM(B{first}:B{last})", font=BOLD_FONT, align=RIGHT,
               number_format=CURRENCY_FMT, border=BORDER_ALL, fill=ROSE_FILL)

    note = ("Avalanche (highest APR first) saves the most interest; snowball "
            "(smallest balance first) gives faster wins. Pick the one you’ll stick with.")
    ws.merge_cells(f"A{tr+2}:I{tr+3}")
    style_cell(ws, f"A{tr+2}", note, font=MUTED_FONT, align=WRAP)

    protect(ws)
    return ws


# =========================================================================== #
#  Free-text sheets
# =========================================================================== #
def build_freetext(wb, name, subtitle, prompts):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 60
    title_block(ws, name, subtitle, span="B1:C1")
    r = 4
    for prompt in prompts:
        if prompt.startswith("##"):
            style_cell(ws, f"B{r}", prompt[2:].strip(), font=HEAD_FONT)
            ws.merge_cells(f"B{r}:C{r}")
            ws.row_dimensions[r].height = 24
        else:
            style_cell(ws, f"B{r}", prompt, font=LABEL_FONT, align=WRAP)
            ans = style_cell(ws, f"C{r}", None, fill=INPUT_FILL, align=WRAP,
                             border=BORDER_ALL, protection=UNLOCK)
            ws.row_dimensions[r].height = 38
        r += 1
    protect(ws)
    return ws


# =========================================================================== #
#  Build it
# =========================================================================== #
def build_workbook(out_path):
    """Build one workbook using the currently-set CURRENCY_SYMBOL / CURRENCY_FMT."""
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    build_readme(wb)
    _, named = build_start_here(wb)
    build_monthly_spending(wb)
    _, nw_ref = build_net_worth(wb)
    named["NetWorth"] = nw_ref
    build_net_worth_log(wb, nw_ref)
    build_freedom_number(wb)
    build_freedom_timeline(wb)
    build_debts(wb)

    build_freetext(wb, "Annual Vision",
                   "Where you want this year to take you.",
                   ["## This year in one sentence",
                    "If this year goes beautifully, what's true by December?",
                    "## Money goals",
                    "Top 3 financial goals for the year:",
                    "What will you stop spending on?",
                    "What will you happily spend more on?",
                    "## Life goals",
                    "Experiences you want to have:",
                    "Who you want to spend time with:",
                    "A skill or habit to build:"])
    build_freetext(wb, "Monthly Review",
                   "A 10-minute check-in, once a month.",
                   ["## Numbers",
                    "Did I hit my savings rate this month?",
                    "Where did I overspend, and why?",
                    "## Reflection",
                    "What money decision am I proud of?",
                    "What felt out of alignment?",
                    "One change for next month:"])
    build_freetext(wb, "Quarterly Review",
                   "Zoom out every 90 days.",
                   ["## Progress",
                    "How much did net worth move this quarter?",
                    "Am I on track for my work-optional year?",
                    "## Course-correct",
                    "What's working that I should do more of?",
                    "What should I stop or change?",
                    "Next quarter's single priority:"])
    build_freetext(wb, "Reflection Prompts",
                   "Open-ended prompts for when you need to think.",
                   ["What does money mean to me, really?",
                    "What's a 'enough' number that would let me breathe?",
                    "What did money stress feel like this year — and what eased it?",
                    "What would I do with an extra free day each week?",
                    "What's one money story from childhood I'm ready to rewrite?",
                    "Who do I want to be generous toward?"])

    # ---- Register named ranges (workbook-level) -------------------------- #
    for name, ref in named.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    wb.save(out_path)
    return len(wb.sheetnames)


def main():
    import os
    global CURRENCY_SYMBOL, CURRENCY_FMT

    # USD is the flagship file at the top level; the rest go in variants/.
    variants_dir = "variants"
    os.makedirs(variants_dir, exist_ok=True)

    for code, symbol, fmt in CURRENCIES:
        CURRENCY_SYMBOL, CURRENCY_FMT = symbol, fmt
        if code == "USD":
            out = "Intentional Budget Tracker.xlsx"
        else:
            out = os.path.join(variants_dir, f"Intentional Budget Tracker — {code}.xlsx")
        n = build_workbook(out)
        print(f"Wrote {out}  ({code} {symbol}, {n} sheets)")


if __name__ == "__main__":
    main()
